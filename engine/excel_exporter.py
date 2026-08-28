"""
서식형 결과 엑셀 파일 생성기 (Excel Report Exporter)
- 공공기관 세출예산 사업명세서 및 예산관리대장 표준 서식 적용
- 1월 ~ 12월 월별 지출 예산 통계 매트릭스 시트
- 1~4회 추경예산 변동 현황표 시트 추가
- openpyxl을 활용한 스타일링 (헤더 색상, 테두리, 통화 서식, 열 너비 자동 조정)
"""

import os
from typing import Dict, Any, List


class ExcelExporter:
    @classmethod
    def export_report(
        cls,
        simulation_data: Dict[str, Any],
        transactions: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl 라이브러리가 필요합니다. pip install openpyxl")

        wb = openpyxl.Workbook()
        default_sheet = wb.active

        # -------------------------------------------------------------
        # 스타일 정의
        # -------------------------------------------------------------
        font_title = Font(name="맑은 고딕", size=15, bold=True, color="1F497D")
        font_subtitle = Font(name="맑은 고딕", size=10, italic=True, color="595959")
        font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        font_body = Font(name="맑은 고딕", size=10)
        font_bold = Font(name="맑은 고딕", size=10, bold=True)
        font_danger = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
        font_warning = Font(name="맑은 고딕", size=10, bold=True, color="ED7D31")
        font_plus = Font(name="맑은 고딕", size=10, color="0070C0", bold=True)
        font_minus = Font(name="맑은 고딕", size=10, color="C00000", bold=True)

        fill_header = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        fill_month_header = PatternFill(start_color="1E4E79", end_color="1E4E79", fill_type="solid")
        fill_supp_header = PatternFill(start_color="4B2C82", end_color="4B2C82", fill_type="solid")
        fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_account = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        thin = Side(border_style="thin", color="BFBFBF")
        thick_bottom = Side(border_style="medium", color="2F4F4F")
        double_bottom = Side(border_style="double", color="000000")

        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_header = Border(left=thin, right=thin, top=thin, bottom=thick_bottom)
        border_total = Border(left=thin, right=thin, top=thin, bottom=double_bottom)

        num_format_currency = "#,##0"
        num_format_diff = "+#,##0;-#,##0;-"
        num_format_percent = "0.0%"

        # =============================================================
        # 1. 시트: 세목별 예산관리대장 (연말 예측 요약 서식)
        # =============================================================
        ws1 = wb.create_sheet(title="세목별 예산관리대장")
        wb.remove(default_sheet)

        ws1.merge_cells("A1:K1")
        ws1["A1"] = "세출예산 세목별 예산관리대장 및 연말(12.31) 예측 현황"
        ws1["A1"].font = font_title
        ws1["A1"].alignment = align_left

        ws1["A2"] = f"※ 기준: 12월 31일 연말 예측 (잔여 정기지출 개월 수: {simulation_data.get('remaining_months', 0)}개월) | 단위: 원"
        ws1["A2"].font = font_subtitle

        headers1 = [
            "통계목", "세목 (산출기초)", "배정예산액 (A)", "기집행액 (B)", "현재잔액 (A-B)", 
            "현재 집행률", "연말 잔여정기지출", "하반기 집행예정", "12.31 예상지출 (C)", "12.31 예상잔액 (A-C)", "상태판정"
        ]
        
        ws1.row_dimensions[4].height = 28
        for col_idx, h_text in enumerate(headers1, start=1):
            cell = ws1.cell(row=4, column=col_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_header

        current_row = 5
        items = simulation_data.get("items", [])
        current_acc = None

        for item in items:
            acc = item["account"]
            if acc != current_acc:
                current_acc = acc
                ws1.cell(row=current_row, column=1, value=acc).font = font_bold
                for c in range(1, len(headers1) + 1):
                    ws1.cell(row=current_row, column=c).fill = fill_account
                    ws1.cell(row=current_row, column=c).border = border_all
                current_row += 1

            ws1.cell(row=current_row, column=1, value=item["account"]).alignment = align_left
            ws1.cell(row=current_row, column=2, value=item["sub_account"]).alignment = align_left
            
            c3 = ws1.cell(row=current_row, column=3, value=item["budget"])
            c4 = ws1.cell(row=current_row, column=4, value=item["actual_spent"])
            c5 = ws1.cell(row=current_row, column=5, value=item["current_balance"])
            c6 = ws1.cell(row=current_row, column=6, value=item["exec_rate"] / 100.0)
            c7 = ws1.cell(row=current_row, column=7, value=item["remaining_recurring"])
            c8 = ws1.cell(row=current_row, column=8, value=item["scheduled_spent"])
            c9 = ws1.cell(row=current_row, column=9, value=item["forecast_total_spent"])
            c10 = ws1.cell(row=current_row, column=10, value=item["forecast_balance"])
            c11 = ws1.cell(row=current_row, column=11, value=item["status"])

            for c in [c3, c4, c5, c7, c8, c9, c10]:
                c.number_format = num_format_currency
                c.alignment = align_right
            c6.number_format = num_format_percent
            c6.alignment = align_right
            c11.alignment = align_center

            if "초과" in item["status"]:
                c10.font = font_danger
                c11.font = font_danger
            elif "불용" in item["status"]:
                c11.font = font_warning

            for c_idx in range(1, len(headers1) + 1):
                cell = ws1.cell(row=current_row, column=c_idx)
                if not cell.font or cell.font.color.rgb not in ("C00000", "ED7D31"):
                    cell.font = font_body
                cell.border = border_all

            current_row += 1

        # 총계 행
        ws1.row_dimensions[current_row].height = 24
        ws1.cell(row=current_row, column=1, value="합 계").alignment = align_center
        ws1.cell(row=current_row, column=2, value="전체 세목 총괄").alignment = align_center
        t_b = ws1.cell(row=current_row, column=3, value=simulation_data.get("total_budget", 0))
        t_s = ws1.cell(row=current_row, column=4, value=simulation_data.get("total_spent", 0))
        t_cb = ws1.cell(row=current_row, column=5, value=simulation_data.get("current_balance", 0))
        t_er = ws1.cell(row=current_row, column=6, value=simulation_data.get("overall_exec_rate", 0) / 100.0)
        t_fs = ws1.cell(row=current_row, column=9, value=simulation_data.get("total_forecast_spent", 0))
        t_fb = ws1.cell(row=current_row, column=10, value=simulation_data.get("total_forecast_balance", 0))
        t_st = ws1.cell(row=current_row, column=11, value=f"최종 {simulation_data.get('overall_forecast_exec_rate', 0)}%")

        for c in [t_b, t_s, t_cb, t_fs, t_fb]:
            c.number_format = num_format_currency
            c.alignment = align_right
        t_er.number_format = num_format_percent
        t_er.alignment = align_right
        t_st.alignment = align_center

        for c_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=current_row, column=c_idx)
            cell.font = font_bold
            cell.fill = fill_total
            cell.border = border_total

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len * 1.5, 13)

        # =============================================================
        # 2. 시트: 📊 추경 예산 변동 현황표 (1~4회 추경 비교)
        # =============================================================
        supp_data = simulation_data.get("supplementary_matrix", {})
        if supp_data and supp_data.get("rows"):
            ws_s = wb.create_sheet(title="추경예산 변동현황(1~4회)")
            ws_s.merge_cells("A1:L1")
            ws_s["A1"] = "세출예산 세목별 1~4회 추가경정예산 변동 이력 및 현황표"
            ws_s["A1"].font = font_title
            ws_s["A1"].alignment = align_left

            ws_s["A2"] = "※ 당초 본예산 대비 1~4회 추경 증감 내역 및 최종 확정예산 대비 집행 현황 | 단위: 원"
            ws_s["A2"].font = font_subtitle

            headers_s = [
                "통계목", "세목 (산출기초)", "당초 본예산",
                "1회 추경(±)", "2회 추경(±)", "3회 추경(±)", "4회 추경(±)",
                "최종 확정예산", "기집행액", "현재잔액", "집행률", "증감사유 및 내역"
            ]

            ws_s.row_dimensions[4].height = 28
            for col_idx, h_text in enumerate(headers_s, start=1):
                cell = ws_s.cell(row=4, column=col_idx, value=h_text)
                cell.font = font_header
                cell.fill = fill_supp_header
                cell.alignment = align_center
                cell.border = border_header

            s_rows = supp_data.get("rows", [])
            s_idx = 5
            curr_s_acc = None

            for sr in s_rows:
                acc = sr["account"]
                if acc != curr_s_acc:
                    curr_s_acc = acc
                    ws_s.cell(row=s_idx, column=1, value=acc).font = font_bold
                    for c in range(1, len(headers_s) + 1):
                        ws_s.cell(row=s_idx, column=c).fill = fill_account
                        ws_s.cell(row=s_idx, column=c).border = border_all
                    s_idx += 1

                ws_s.cell(row=s_idx, column=1, value=sr["account"]).alignment = align_left
                ws_s.cell(row=s_idx, column=2, value=sr["sub_account"]).alignment = align_left
                
                c_base = ws_s.cell(row=s_idx, column=3, value=sr["base_budget"])
                c_r1 = ws_s.cell(row=s_idx, column=4, value=sr["r1"])
                c_r2 = ws_s.cell(row=s_idx, column=5, value=sr["r2"])
                c_r3 = ws_s.cell(row=s_idx, column=6, value=sr["r3"])
                c_r4 = ws_s.cell(row=s_idx, column=7, value=sr["r4"])
                c_fin = ws_s.cell(row=s_idx, column=8, value=sr["final_budget"])
                c_sp = ws_s.cell(row=s_idx, column=9, value=sr["spent"])
                c_bl = ws_s.cell(row=s_idx, column=10, value=sr["balance"])
                c_rt = ws_s.cell(row=s_idx, column=11, value=sr["exec_rate"] / 100.0)
                c_rs = ws_s.cell(row=s_idx, column=12, value=sr["reason"])

                for c in [c_base, c_fin, c_sp, c_bl]:
                    c.number_format = num_format_currency
                    c.alignment = align_right

                for c_r, val in [(c_r1, sr["r1"]), (c_r2, sr["r2"]), (c_r3, sr["r3"]), (c_r4, sr["r4"])]:
                    c_r.number_format = num_format_diff
                    c_r.alignment = align_right
                    if val > 0:
                        c_r.font = font_plus
                    elif val < 0:
                        c_r.font = font_minus

                c_rt.number_format = num_format_percent
                c_rt.alignment = align_right
                c_rs.alignment = align_left

                for c_idx in range(1, len(headers_s) + 1):
                    cell = ws_s.cell(row=s_idx, column=c_idx)
                    if not cell.font:
                        cell.font = font_body
                    cell.border = border_all

                s_idx += 1

            # 추경 총계 행
            ws_s.row_dimensions[s_idx].height = 24
            ws_s.cell(row=s_idx, column=1, value="합 계").alignment = align_center
            ws_s.cell(row=s_idx, column=2, value="전체 추경 총괄").alignment = align_center
            
            c_tb = ws_s.cell(row=s_idx, column=3, value=supp_data.get("total_base", 0))
            c_tr1 = ws_s.cell(row=s_idx, column=4, value=supp_data.get("total_r1", 0))
            c_tr2 = ws_s.cell(row=s_idx, column=5, value=supp_data.get("total_r2", 0))
            c_tr3 = ws_s.cell(row=s_idx, column=6, value=supp_data.get("total_r3", 0))
            c_tr4 = ws_s.cell(row=s_idx, column=7, value=supp_data.get("total_r4", 0))
            c_tfin = ws_s.cell(row=s_idx, column=8, value=supp_data.get("total_final", 0))
            c_tsp = ws_s.cell(row=s_idx, column=9, value=supp_data.get("total_spent", 0))
            c_tbl = ws_s.cell(row=s_idx, column=10, value=supp_data.get("total_balance", 0))
            c_trt = ws_s.cell(row=s_idx, column=11, value=supp_data.get("overall_exec_rate", 0) / 100.0)
            ws_s.cell(row=s_idx, column=12, value="-").alignment = align_center

            for c in [c_tb, c_tfin, c_tsp, c_tbl]:
                c.number_format = num_format_currency
                c.alignment = align_right
            for c in [c_tr1, c_tr2, c_tr3, c_tr4]:
                c.number_format = num_format_diff
                c.alignment = align_right
            c_trt.number_format = num_format_percent
            c_trt.alignment = align_right

            for c_idx in range(1, len(headers_s) + 1):
                cell = ws_s.cell(row=s_idx, column=c_idx)
                cell.font = font_bold
                cell.fill = fill_total
                cell.border = border_total

            for col in ws_s.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_s.column_dimensions[col_letter].width = max(max_len * 1.4, 12)

        # =============================================================
        # 3. 시트: 📅 월별 지출 예산 통계표 (1~12월 매트릭스)
        # =============================================================
        ws_m = wb.create_sheet(title="월별 지출현황(1~12월)")
        
        ws_m.merge_cells("A1:R1")
        ws_m["A1"] = "통계목 및 세목별 월별(1월~12월) 지출 통계 현황표"
        ws_m["A1"].font = font_title
        ws_m["A1"].alignment = align_left

        ws_m["A2"] = "※ e호조 지출내역 기준 월별 집행 추이 및 잔액 분석 | 단위: 원"
        ws_m["A2"].font = font_subtitle

        headers_m = [
            "통계목", "세목 (산출기초)", "배정예산액",
            "1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월",
            "누적집행액", "현재잔액", "집행률"
        ]

        ws_m.row_dimensions[4].height = 28
        for col_idx, h_text in enumerate(headers_m, start=1):
            cell = ws_m.cell(row=4, column=col_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_month_header
            cell.alignment = align_center
            cell.border = border_header

        m_data = simulation_data.get("monthly_matrix", {})
        m_rows = m_data.get("rows", [])
        r_idx = 5
        curr_m_acc = None

        for row in m_rows:
            acc = row["account"]
            if acc != curr_m_acc:
                curr_m_acc = acc
                ws_m.cell(row=r_idx, column=1, value=acc).font = font_bold
                for c in range(1, len(headers_m) + 1):
                    ws_m.cell(row=r_idx, column=c).fill = fill_account
                    ws_m.cell(row=r_idx, column=c).border = border_all
                r_idx += 1

            ws_m.cell(row=r_idx, column=1, value=row["account"]).alignment = align_left
            ws_m.cell(row=r_idx, column=2, value=row["sub_account"]).alignment = align_left
            
            c_bg = ws_m.cell(row=r_idx, column=3, value=row["budget"])
            c_bg.number_format = num_format_currency
            c_bg.alignment = align_right

            for m_i, m_val in enumerate(row["months"], start=1):
                c_m = ws_m.cell(row=r_idx, column=3 + m_i, value=m_val)
                c_m.number_format = num_format_currency
                c_m.alignment = align_right

            c_tot = ws_m.cell(row=r_idx, column=16, value=row["total_spent"])
            c_bal = ws_m.cell(row=r_idx, column=17, value=row["balance"])
            c_rate = ws_m.cell(row=r_idx, column=18, value=row["exec_rate"] / 100.0)

            c_tot.number_format = num_format_currency
            c_bal.number_format = num_format_currency
            c_rate.number_format = num_format_percent
            c_tot.alignment = align_right
            c_bal.alignment = align_right
            c_rate.alignment = align_right

            for c_idx in range(1, len(headers_m) + 1):
                cell = ws_m.cell(row=r_idx, column=c_idx)
                if not cell.font:
                    cell.font = font_body
                cell.border = border_all

            r_idx += 1

        # 월별 총계 행
        ws_m.row_dimensions[r_idx].height = 24
        ws_m.cell(row=r_idx, column=1, value="합 계").alignment = align_center
        ws_m.cell(row=r_idx, column=2, value="전체 월별 총괄").alignment = align_center
        
        c_t_bg = ws_m.cell(row=r_idx, column=3, value=m_data.get("total_budget", 0))
        c_t_bg.number_format = num_format_currency
        c_t_bg.alignment = align_right

        m_totals = m_data.get("monthly_totals", [0] * 12)
        for m_i, m_val in enumerate(m_totals, start=1):
            c_tm = ws_m.cell(row=r_idx, column=3 + m_i, value=m_val)
            c_tm.number_format = num_format_currency
            c_tm.alignment = align_right

        c_t_tot = ws_m.cell(row=r_idx, column=16, value=m_data.get("total_spent", 0))
        c_t_bal = ws_m.cell(row=r_idx, column=17, value=m_data.get("total_balance", 0))
        c_t_rate = ws_m.cell(row=r_idx, column=18, value=m_data.get("overall_exec_rate", 0) / 100.0)

        c_t_tot.number_format = num_format_currency
        c_t_bal.number_format = num_format_currency
        c_t_rate.number_format = num_format_percent
        c_t_tot.alignment = align_right
        c_t_bal.alignment = align_right
        c_t_rate.alignment = align_right

        for c_idx in range(1, len(headers_m) + 1):
            cell = ws_m.cell(row=r_idx, column=c_idx)
            cell.font = font_bold
            cell.fill = fill_total
            cell.border = border_total

        for col in ws_m.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_m.column_dimensions[col_letter].width = max(max_len * 1.3, 11)

        # =============================================================
        # 4. 시트: 지출 상세 내역
        # =============================================================
        ws2 = wb.create_sheet(title="e호조 수집 지출내역")
        headers2 = ["연번", "결의일자", "통계목", "자동분류 세목", "지출 적요 (온나라 기안명)", "채권자 (지급처)", "지출액 (원)", "적용 키워드 규칙"]
        
        ws2.row_dimensions[1].height = 25
        for col_idx, h_text in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_header

        for idx, tx in enumerate(transactions, start=1):
            r = idx + 1
            ws2.cell(row=r, column=1, value=idx).alignment = align_center
            ws2.cell(row=r, column=2, value=tx.get("date", "")).alignment = align_center
            ws2.cell(row=r, column=3, value=tx.get("account", "")).alignment = align_left
            
            sub_cell = ws2.cell(row=r, column=4, value=tx.get("sub_account", "미분류"))
            sub_cell.alignment = align_left
            if tx.get("sub_account") == "미분류":
                sub_cell.font = font_warning

            ws2.cell(row=r, column=5, value=tx.get("summary", "")).alignment = align_left
            ws2.cell(row=r, column=6, value=tx.get("vendor", "")).alignment = align_left
            
            amt_cell = ws2.cell(row=r, column=7, value=int(tx.get("amount", 0)))
            amt_cell.number_format = num_format_currency
            amt_cell.alignment = align_right
            
            ws2.cell(row=r, column=8, value=tx.get("rule_matched", "")).alignment = align_left

            for c_idx in range(1, len(headers2) + 1):
                cell = ws2.cell(row=r, column=c_idx)
                if not cell.font:
                    cell.font = font_body
                cell.border = border_all

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len * 1.4, 12)

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path
